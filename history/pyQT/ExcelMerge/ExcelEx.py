import os

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


def copy_data_to_template(fileList, template_file, output_file):
    try:
        # 加载模板文件
        template_workbook = load_workbook(template_file)
        template_sheet = template_workbook.active

        for file in fileList:
            source_file = file
            try:
                # 这里使用 data_only=True 来获取公式计算结果
                source_workbook = load_workbook(source_file, data_only=True)
                source_sheet = source_workbook.active
                print("current file:" + source_file)
                # 输出源文件基本信息
                print(f"源文件活动工作表名称: {source_sheet.title}")
                print(f"源文件活动工作表最大行数: {source_sheet.max_row}")
                print(f"源文件活动工作表最大列数: {source_sheet.max_column}")

                has_error = False
                # 逐行复制数据，从第二行开始
                for row_index, row in enumerate(source_sheet.iter_rows(min_row=2), start=2):
                    for col_index, cell in enumerate(row, start=1):
                        if cell.data_type == 'e':  # 检查是否为错误值
                            print(f"发现问题表格: {source_file}，第 {row_index} 行，第 {get_column_letter(col_index)} 列存在错误单元格: {cell.value}")
                            has_error = True
                            break
                    if has_error:
                        break

                if not has_error:
                    # 逐行复制数据
                    for row in source_sheet.iter_rows(min_row=2):
                        # 找到模板表的下一行
                        next_row = template_sheet.max_row + 1
                        for col_index, cell in enumerate(row, start=1):
                            value = cell.value
                            # 复制值并保留格式
                            template_cell = template_sheet.cell(row=next_row, column=col_index)
                            template_cell.value = value

            except Exception as e:
                print(f"处理文件 {source_file} 时发生错误: {e}")

        # 保存到输出文件
        template_workbook.save(output_file)
        print(f"数据已复制到 '{output_file}'")

        return True

    except Exception as e:
        print(f"发生错误: {e}")
        return False

def datasheet_copy(fileList, output_file):


    # 执行数据复制
    copy_data_to_template(fileList, template_file, output_file)

def datasheet_copy(fileList, output_file):
    # 文件列表
    print("开始进行表格复制，所有表格路径如下" + fileList[0])
    # 查看列表中的所有文件是否都存在
    template_file = './src/origin.xlsx'  # 模板文件路径
    non_existent_files = []
    for file in fileList:
        if not os.path.exists(file):
            non_existent_files.append(file)
    if non_existent_files:
        print(f"以下文件不存在: {non_existent_files}")
        return
    if not os.path.exists(template_file):
        print(f"模板文件 {template_file} 不存在")
        return
    # 执行数据复制
    if copy_data_to_template(fileList, template_file, output_file):
        return True
    else:
        return False


if __name__ == '__main__':
    # 文件列表
    fileList = ['./test/1111.xlsx', './test/2222.xlsx', './test/3333.xlsx', './test/11111.xlsx', './test/11112.xlsx',
                './test/22222.xlsx', './test/33333.xlsx']
    template_file = './src/origin1.xlsx'  # 模板文件路径
    output_file = 'D:/GitFiles/pythonFiles/pyQT/ExcelMerge/updated_template.xlsx'  # 输出文件路径

    # 执行数据复制
    copy_data_to_template(fileList, template_file, output_file)
